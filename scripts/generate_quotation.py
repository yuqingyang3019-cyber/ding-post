#!/usr/bin/env python3
"""生成项目报价单 Excel 文件。"""

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Sequence, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT = Path(__file__).resolve().parent.parent / "报价单.xlsx"

# 配色
COLOR_PRIMARY = "1F4E79"
COLOR_PRIMARY_LIGHT = "2E75B6"
COLOR_HEADER_BG = "D6E4F0"
COLOR_ROW_ALT = "F5F8FC"
COLOR_SUMMARY = "FFF4E5"
COLOR_DISCOUNT = "E8F4EA"
COLOR_BORDER = "B4C6D8"
COLOR_TEXT_MUTED = "5A6A7A"
COLOR_WHITE = "FFFFFF"

FONT_TITLE = "微软雅黑"
FONT_BODY = "微软雅黑"

HEADERS = ("序号", "需求清单", "报价（人民币/元）")
Price = Union[int, float, str]

DEFAULT_NOTE = (
    "备注：以上报价含开发与部署费用；阿里云服务器为年费，可按项目共用分摊。最终价格以双方确认为准。"
)


@dataclass
class QuotationSheet:
    sheet_name: str
    title: str
    rows: Sequence[tuple[str, Price]]
    summary_rows: Sequence[tuple[str, Price]]
    note: str = DEFAULT_NOTE
    row_heights: dict[int, float] = field(default_factory=dict)


SHEETS = [
    QuotationSheet(
        sheet_name="价格采集",
        title="价格采集项目报价单",
        rows=[
            ("后端支持OCR图片处理，用户确认后存储到数据库。", 2000),
            ("后端支持数据计算，例如涨幅均差等。支持历史记录查询", 2000),
            (
                '用户上传图片处理成功后，机器人同步在群内更新"某某已上传当日价格"消息',
                2000,
            ),
            (
                "前端支持用户上传图片，并用户确认。前端支持图表展示（展示内容根据甲方要求改）。前端支持历史记录查看。",
                4000,
            ),
            ("集成该应用到钉钉中，能够在工作台找到企业内部应用。", 4000),
            ("阿里云服务器采购（可与其他项目共用）", "2190.2 / 年"),
        ],
        summary_rows=[
            ("总价", 16190.2),
            (
                "优惠价（钉钉集成已在其他项目付费，阿里云服务器不参与优惠）",
                "¥9,000 + ¥2,190 = ¥11,190",
            ),
        ],
    ),
    QuotationSheet(
        sheet_name="用友ERP",
        title="用友ERP对接项目报价单",
        rows=[
            (
                "前端支持用户按天、按月、按季度查询用友信息，并按照甲方要求进行修改"
                "（不支持走势图，如需走势图需额外增加数据库）",
                2000,
            ),
            (
                "后端支持接入用友 ERP 五个单据信息：采购合同、采购订单、采购到货、"
                "付款申请单、采购发票。其中采购发票单据爬取创建人、采购员、"
                "含税金额、创建日期、单据日期。",
                4000,
            ),
            (
                "后端依据甲方要求，在以上五个单据的数据基础上进行数学公式规则开发，"
                "例如根据采购发票计算每个人的执行数量和执行金额等。",
                2000,
            ),
            ("钉钉集成（含登录鉴权）", 4000),
            ("阿里云服务采购", "2190.2 / 年"),
        ],
        summary_rows=[
            ("总价", 10190.2),
            (
                "优惠价（钉钉集成和阿里云服务器采购已在其他项目付费）",
                6300,
            ),
        ],
        row_heights={5: 48, 6: 52, 7: 44},
    ),
]


def _side(style: str = "thin", color: str = COLOR_BORDER) -> Side:
    return Side(style=style, color=color)


def border_all() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def set_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    *,
    font=None,
    alignment=None,
    border=None,
    fill_color=None,
    number_format=None,
):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill_color:
        cell.fill = fill(fill_color)
    if number_format:
        cell.number_format = number_format
    return cell


def build_quotation_sheet(ws: Worksheet, config: QuotationSheet, today: str) -> None:
    ws.title = config.sheet_name
    ws.sheet_view.showGridLines = False

    border = border_all()
    data_start = 5

    ws.merge_cells("A1:C1")
    set_cell(
        ws, 1, 1, config.title,
        font=Font(name=FONT_TITLE, bold=True, size=18, color=COLOR_WHITE),
        alignment=Alignment(horizontal="center", vertical="center"),
        fill_color=COLOR_PRIMARY,
    )
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:C2")
    set_cell(
        ws, 2, 1, f"报价日期：{today}　　　币种：人民币（CNY）",
        font=Font(name=FONT_BODY, size=10, color=COLOR_TEXT_MUTED),
        alignment=Alignment(horizontal="center", vertical="center"),
        fill_color="F0F4F8",
    )
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 8

    header_font = Font(name=FONT_BODY, bold=True, size=11, color=COLOR_PRIMARY)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, title in enumerate(HEADERS, start=1):
        set_cell(
            ws, 4, col, title,
            font=header_font,
            alignment=header_align,
            border=border,
            fill_color=COLOR_HEADER_BG,
        )
    ws.row_dimensions[4].height = 32

    body_font = Font(name=FONT_BODY, size=10, color="333333")
    idx_font = Font(name=FONT_BODY, size=10, color=COLOR_TEXT_MUTED)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center", indent=1)
    money_fmt = '#,##0.00" 元"'

    row_idx = data_start
    for i, (requirement, price) in enumerate(config.rows, start=1):
        bg = COLOR_ROW_ALT if i % 2 == 0 else COLOR_WHITE
        set_cell(ws, row_idx, 1, i, font=idx_font, alignment=center, border=border, fill_color=bg)
        set_cell(ws, row_idx, 2, requirement, font=body_font, alignment=left_wrap, border=border, fill_color=bg)
        if isinstance(price, (int, float)):
            set_cell(
                ws, row_idx, 3, price,
                font=Font(name=FONT_BODY, size=10, color=COLOR_PRIMARY),
                alignment=right_align,
                border=border,
                fill_color=bg,
                number_format=money_fmt,
            )
        else:
            set_cell(
                ws, row_idx, 3, price,
                font=body_font,
                alignment=right_align,
                border=border,
                fill_color=bg,
            )
        ws.row_dimensions[row_idx].height = config.row_heights.get(row_idx, 40)
        row_idx += 1

    summary_label_font = Font(name=FONT_BODY, bold=True, size=11, color=COLOR_PRIMARY)
    summary_price_font = Font(name=FONT_BODY, bold=True, size=12, color=COLOR_PRIMARY_LIGHT)
    summary_bg_colors = (COLOR_SUMMARY, COLOR_DISCOUNT)

    for j, (label, price) in enumerate(config.summary_rows):
        bg_color = summary_bg_colors[j] if j < len(summary_bg_colors) else COLOR_SUMMARY
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        set_cell(
            ws, row_idx, 1, label,
            font=summary_label_font,
            alignment=left_wrap,
            border=border,
            fill_color=bg_color,
        )
        set_cell(ws, row_idx, 2, None, border=border, fill_color=bg_color)

        if isinstance(price, (int, float)):
            set_cell(
                ws, row_idx, 3, price,
                font=summary_price_font,
                alignment=right_align,
                border=border,
                fill_color=bg_color,
                number_format=money_fmt,
            )
        else:
            set_cell(
                ws, row_idx, 3, price,
                font=summary_price_font,
                alignment=right_align,
                border=border,
                fill_color=bg_color,
            )
        ws.row_dimensions[row_idx].height = 36
        row_idx += 1

    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
    set_cell(
        ws, row_idx, 1, config.note,
        font=Font(name=FONT_BODY, size=9, color=COLOR_TEXT_MUTED, italic=True),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1),
    )
    ws.row_dimensions[row_idx].height = 28

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 24

    ws.freeze_panes = "A5"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def main() -> None:
    wb = Workbook()
    today = date.today().strftime("%Y年%m月%d日")

    for i, config in enumerate(SHEETS):
        ws = wb.active if i == 0 else wb.create_sheet()
        build_quotation_sheet(ws, config, today)

    wb.save(OUTPUT)
    print(f"已生成: {OUTPUT}（共 {len(SHEETS)} 个工作表）")


if __name__ == "__main__":
    main()
